from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from pathlib import Path

from ..models import DatasetKind, ValidationSeverity

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency installed at runtime
    openpyxl = None

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency installed at runtime
    xlrd = None


@dataclass
class ParsedRow:
    sheet_name: str
    row_index: int
    raw_json: str


@dataclass
class ParsedIssue:
    severity: ValidationSeverity
    rule_code: str
    message: str
    sheet_name: str | None = None
    row_index: int | None = None


@dataclass
class ParsedWorkbook:
    dataset_kind: DatasetKind
    total_sheets: int
    rows: list[ParsedRow]
    issues: list[ParsedIssue]


def parse_file(filename: str, payload: bytes) -> ParsedWorkbook:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(payload)
    if suffix == ".xlsx":
        return _parse_xlsx(payload)
    if suffix == ".xls":
        return _parse_xls(payload)
    return ParsedWorkbook(
        dataset_kind=DatasetKind.unknown,
        total_sheets=0,
        rows=[],
        issues=[
            ParsedIssue(
                severity=ValidationSeverity.error,
                rule_code="UNSUPPORTED_EXTENSION",
                message=f"Unsupported file type: {suffix or 'no extension'}",
            )
        ],
    )


def _detect_dataset_kind(sheet_names: list[str]) -> DatasetKind:
    normalized = [name.strip().lower() for name in sheet_names]
    if any("." in name and len(name) <= 5 for name in normalized):
        return DatasetKind.daily_summary
    if any("баланс" in name or "тех" in name for name in normalized):
        return DatasetKind.technical_balance
    return DatasetKind.unknown


def _parse_csv(payload: bytes) -> ParsedWorkbook:
    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [
        ParsedRow("csv", index, json.dumps(cells, ensure_ascii=False))
        for index, cells in enumerate(reader, start=1)
    ]
    issues = _basic_issues(["csv"], rows)
    return ParsedWorkbook(
        dataset_kind=DatasetKind.unknown,
        total_sheets=1,
        rows=rows,
        issues=issues,
    )


def _parse_xlsx(payload: bytes) -> ParsedWorkbook:
    if openpyxl is None:
        return _missing_dependency("openpyxl")
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False, read_only=True)
    rows: list[ParsedRow] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            rows.append(
                ParsedRow(sheet_name, row_index, json.dumps(list(row), ensure_ascii=False, default=str))
            )
    issues = _basic_issues(workbook.sheetnames, rows)
    return ParsedWorkbook(
        dataset_kind=_detect_dataset_kind(workbook.sheetnames),
        total_sheets=len(workbook.sheetnames),
        rows=rows,
        issues=issues,
    )


def _parse_xls(payload: bytes) -> ParsedWorkbook:
    if xlrd is None:
        return _missing_dependency("xlrd")
    workbook = xlrd.open_workbook(file_contents=payload)
    rows: list[ParsedRow] = []
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            rows.append(
                ParsedRow(
                    sheet.name,
                    row_index + 1,
                    json.dumps(sheet.row_values(row_index), ensure_ascii=False, default=str),
                )
            )
    sheet_names = workbook.sheet_names()
    issues = _basic_issues(sheet_names, rows)
    return ParsedWorkbook(
        dataset_kind=_detect_dataset_kind(sheet_names),
        total_sheets=len(sheet_names),
        rows=rows,
        issues=issues,
    )


def _basic_issues(sheet_names: list[str], rows: list[ParsedRow]) -> list[ParsedIssue]:
    issues: list[ParsedIssue] = []
    if not rows:
        issues.append(
            ParsedIssue(
                severity=ValidationSeverity.error,
                rule_code="EMPTY_FILE",
                message="Uploaded file contains no readable rows.",
            )
        )
    if not sheet_names:
        issues.append(
            ParsedIssue(
                severity=ValidationSeverity.error,
                rule_code="NO_SHEETS",
                message="Workbook contains no sheets.",
            )
        )
    return issues


def _missing_dependency(name: str) -> ParsedWorkbook:
    return ParsedWorkbook(
        dataset_kind=DatasetKind.unknown,
        total_sheets=0,
        rows=[],
        issues=[
            ParsedIssue(
                severity=ValidationSeverity.error,
                rule_code="MISSING_DEPENDENCY",
                message=f"Runtime dependency is not installed: {name}",
            )
        ],
    )
